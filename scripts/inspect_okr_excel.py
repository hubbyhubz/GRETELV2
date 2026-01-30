import argparse

import openpyxl


def norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace("\n", " ").strip()
    return str(v)


def find_header_row(ws, max_rows: int, max_cols: int):
    keywords = {
        "objective",
        "key result",
        "kr",
        "target",
        "target type",
        "initiatives",
        "weight",
        "weightage",
        "metric",
        "achieved",
        "status",
        "start",
        "end",
    }
    best = None
    for r in range(1, max_rows + 1):
        row = [norm(ws.cell(r, c).value).lower() for c in range(1, max_cols + 1)]
        hits = 0
        for cell in row:
            for k in keywords:
                if k in cell and cell:
                    hits += 1
                    break
        if hits:
            if best is None or hits > best[0]:
                best = (hits, r, row)
    return None if best is None else best[1]


def dump_sheet(ws, max_rows: int, max_cols: int):
    header_row = find_header_row(ws, max_rows=min(max_rows, 80), max_cols=max_cols)
    if header_row:
        start_row = max(1, header_row - 2)
    else:
        start_row = 1

    rows = []
    for r in range(start_row, start_row + max_rows):
        out = [norm(ws.cell(r, c).value) for c in range(1, max_cols + 1)]
        if any(x for x in out):
            rows.append((r, out))
    return header_row, rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--sheet")
    ap.add_argument("--rows", type=int, default=35)
    ap.add_argument("--cols", type=int, default=25)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path, data_only=True)
    sheets = wb.sheetnames
    print("SHEETS:\n" + "\n".join(f"- {s}" for s in sheets))

    target_sheets = [args.sheet] if args.sheet else sheets[:1]
    for name in target_sheets:
        ws = wb[name]
        header_row, rows = dump_sheet(ws, max_rows=args.rows, max_cols=args.cols)
        print("\n" + ("=" * 80))
        print(f"SHEET: {name}")
        print(f"HEADER_ROW_GUESS: {header_row or 'n/a'}")
        for r, out in rows:
            print(str(r) + "\t" + "\t".join(out))


if __name__ == "__main__":
    main()

